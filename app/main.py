"""KH07 Welfare — FastAPI + Jinja2 + HTMX."""
import sys, io, os, json, logging
from pathlib import Path
from datetime import date, datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy import select, func, desc, String, extract
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.database import init_db, get_db, count_members, count_causes, sum_contributions, count_contributions, member_total_and_count
from app.models import Member, ContributionCause, Contribution, Disbursement, User, MpesaConfig, MpesaTransaction
from app.auth import require_auth, require_admin, verify_password, create_session, SESSION_COOKIE, SESSION_MAX_AGE, get_session_user

from jinja2 import Environment, FileSystemLoader

logger = logging.getLogger("kh07")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")

templates_dir = Path(__file__).parent / "templates"
env = Environment(loader=FileSystemLoader(str(templates_dir)))


def render(name: str, **ctx) -> HTMLResponse:
    template = env.get_template(name)
    return HTMLResponse(template.render(**ctx))


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    yield


app = FastAPI(title="KH07 Welfare", lifespan=lifespan)

static_dir = Path(__file__).parent / "static"
if static_dir.exists():
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

# ── Register routes directly on app (avoids FastAPI 0.139 include_router bug) ──

# Debug
@app.get("/ping")
async def ping():
    return {"status": "ok"}


@app.get("/check-auth")
async def check_auth(request: Request):
    user = get_session_user(request)
    return JSONResponse({
        "authenticated": user is not None,
        "user": user,
        "cookie_present": SESSION_COOKIE in request.cookies,
    })


# Auth
@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = ""):
    if get_session_user(request):
        return RedirectResponse(url="/alumni", status_code=302)
    return render("login.html", error=error)


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.username == username, User.is_active == True))
    user = result.scalar_one_or_none()
    if user and verify_password(password, user.password_hash):
        token = create_session(username, user.role)
        resp = RedirectResponse(url="/alumni", status_code=302)
        resp.set_cookie(key=SESSION_COOKIE, value=token, max_age=int(SESSION_MAX_AGE.total_seconds()), httponly=True, samesite="lax")
        return resp
    return render("login.html", error="Invalid username or password")


@app.get("/logout")
async def logout(request: Request):
    resp = RedirectResponse(url="/login", status_code=302)
    resp.delete_cookie(SESSION_COOKIE)
    return resp


# Landing page (public)
@app.get("/", response_class=HTMLResponse)
async def landing_page(request: Request, db: AsyncSession = Depends(get_db)):
    total_members = (await db.execute(select(func.count(Member.id)))).scalar() or 0
    total_causes = (await db.execute(select(func.count(ContributionCause.id)).where(ContributionCause.is_active == True))).scalar() or 0
    total_collected = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)))).scalar() or 0)
    total_contributions = (await db.execute(select(func.count(Contribution.id)))).scalar() or 0
    return render("landing.html", request=request,
                  total_members=total_members, total_causes=total_causes,
                  total_collected=total_collected, total_contributions=total_contributions)


# ── Member self-service portal ──
@app.get("/self-service", response_class=HTMLResponse)
async def portal_page(request: Request, db: AsyncSession = Depends(get_db)):
    return render("portal.html", request=request, member=None, contributions=None, total=0, cause_totals={})


@app.post("/self-service/lookup")
async def portal_lookup(request: Request, query: str = Form(""), db: AsyncSession = Depends(get_db)):
    q = query.strip()
    if not q:
        return HTMLResponse('<div class="alert alert-danger">Enter a name, phone number, or member number</div>')
    
    # Try exact matches first
    member = None
    members = []
    
    # Member number match
    if q.isdigit():
        m = (await db.execute(select(Member).where(Member.member_number == int(q)))).scalar_one_or_none()
        if m: member = m
    
    # Phone exact match
    if not member:
        m = (await db.execute(select(Member).where(Member.phone_number == q))).scalar_one_or_none()
        if m: member = m
    
    # Name exact match
    if not member:
        m = (await db.execute(select(Member).where(Member.name == q))).scalar_one_or_none()
        if m: member = m
    
    # Partial matches
    if not member:
        name_matches = (await db.execute(select(Member).where(Member.name.ilike(f"%{q}%")).limit(8))).scalars().all()
        phone_matches = (await db.execute(select(Member).where(Member.phone_number.contains(q)).limit(8))).scalars().all()
        members = list({m.id: m for m in name_matches + phone_matches}.values())  # deduplicate
        
        if len(members) == 1:
            member = members[0]
            members = []
    
    if not member and not members:
        return HTMLResponse(f'<div class="alert alert-danger">No member found matching <strong>{q}</strong>. Try searching by name, phone number (e.g. 0712...), or member number.</div>')
    
    # Show picker if multiple matches
    if members and len(members) > 1:
        rows = "".join(f'<tr hx-post="/self-service/lookup" hx-target="#portal-result" hx-swap="innerHTML" hx-vals=\'{{"query": "{m.name}"}}\' style="cursor:pointer"><td>{m.member_number}</td><td>{m.name}</td></tr>' for m in members)
        return HTMLResponse(f'''<div class="card"><div class="card-header"><i class="fas fa-users me-2" style="color:var(--warning)"></i>Multiple members found</div>
            <div class="card-body p-0"><table class="table table-hover mb-0"><thead><tr><th class="ps-3">#</th><th>Name</th></tr></thead><tbody>{rows}</tbody></table>
            <div class="p-3 text-center text-muted small">Click the matching member above</div></div></div>''')
    
    if not member and members:
        member = members[0]
    
    # Get contributions
    contribs = await db.execute(
        select(Contribution).where(Contribution.member_id == member.id)
        .options(selectinload(Contribution.cause))
        .order_by(desc(Contribution.date_paid)))
    contributions = contribs.scalars().all()
    total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == member.id))).scalar() or 0)
    cause_totals = {}
    for c in contributions:
        name = c.cause.name
        cause_totals[name] = cause_totals.get(name, 0) + float(c.amount)
    
    # Get active causes for portal display
    active_causes = (await db.execute(
        select(ContributionCause, func.coalesce(func.sum(Contribution.amount), 0).label("raised"))
        .outerjoin(Contribution, ContributionCause.id == Contribution.cause_id)
        .where(ContributionCause.is_active == True)
        .group_by(ContributionCause.id)
        .order_by(ContributionCause.name)
    )).all()
    cause_list = []
    for cause, raised in active_causes:
        target = float(cause.target_amount or 0)
        cause_list.append({
            "name": cause.name,
            "raised": float(raised),
            "target": target,
            "progress": (float(raised) / target * 100) if target > 0 else 0,
        })
    
    return render("portal.html", request=request, member=member, contributions=contributions,
                  total=total, cause_totals=cause_totals, causes=cause_list)


@app.post("/self-service/update-phone")
async def portal_update_phone(request: Request, member_id: int = Form(...), phone: str = Form(""), db: AsyncSession = Depends(get_db)):
    member = await db.get(Member, member_id)
    if not member:
        return HTMLResponse('<div class="alert alert-danger">Member not found</div>')
    member.phone_number = phone.strip()
    await db.commit()
    return HTMLResponse(f'<div class="alert alert-success"><i class="fas fa-check-circle me-1"></i>Phone updated to <strong>{phone.strip()}</strong></div>')


@app.get("/self-service/suggest-cause", response_class=HTMLResponse)
async def portal_suggest_cause(request: Request):
    return render("portal_suggest.html", request=request)


@app.post("/self-service/suggest-cause")


@app.get("/welfare-causes/{cause_id}", response_class=HTMLResponse)
async def cause_detail(cause_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = await db.get(ContributionCause, cause_id)
    if not cause:
        return HTMLResponse("Cause not found", status_code=404)
    
    # Total raised
    total_raised = float((await db.execute(
        select(func.coalesce(func.sum(Contribution.amount), 0))
        .where(Contribution.cause_id == cause_id))).scalar() or 0)
    
    # Total disbursed
    total_disbursed = float((await db.execute(
        select(func.coalesce(func.sum(Disbursement.amount), 0))
        .where(Disbursement.cause_id == cause_id))).scalar() or 0)
    
    # Contributor count
    contributor_count = (await db.execute(
        select(func.count(func.distinct(Contribution.member_id)))
        .where(Contribution.cause_id == cause_id))).scalar() or 0
    
    # All active members with their contribution status
    members = await db.execute(
        select(Member)
        .where(Member.is_active == True)
        .order_by(Member.name))
    members = members.scalars().all()
    
    # Get all contributions for this cause
    contribs = await db.execute(
        select(Contribution)
        .where(Contribution.cause_id == cause_id)
        .options(selectinload(Contribution.member)))
    contribs = contribs.scalars().all()
    contrib_map = {c.member_id: c for c in contribs}
    
    member_status = []
    for m in members:
        c = contrib_map.get(m.id)
        member_status.append({
            "id": m.id,
            "name": m.name,
            "member_number": m.member_number,
            "paid": c is not None,
            "amount": float(c.amount) if c else 0,
            "date_paid": c.date_paid if c else None,
            "payment_method": c.payment_method if c else None,
            "contrib_id": c.id if c else None,
        })
    
    # Stats
    outstanding = max(0, total_raised - total_disbursed)
    target = float(cause.target_amount or 0)
    progress = (total_raised / target * 100) if target > 0 else 0
    total_members = len(members)
    non_contributors = total_members - contributor_count
    
    return render("cause_detail.html", user=user, request=request, cause=cause,
                  total_raised=total_raised, total_disbursed=total_disbursed,
                  outstanding=outstanding, target=target, progress=progress,
                  contributor_count=contributor_count, total_members=total_members,
                  non_contributors=non_contributors, member_status=member_status)

async def portal_suggest_submit(request: Request, name: str = Form(...), reason: str = Form(""), db: AsyncSession = Depends(get_db)):
    cause = ContributionCause(name=f"[Suggestion] {name.strip()}", is_active=False)
    db.add(cause)
    await db.commit()
    return HTMLResponse(f'<div class="alert alert-success"><i class="fas fa-check-circle me-1"></i>Your suggestion for "<strong>{name.strip()}</strong>" has been submitted for admin review.</div>')


# Dashboard (requires auth)
@app.get("/overview", response_class=HTMLResponse)
async def dashboard(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_session_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    # Aggregate stats
    total_members = (await db.execute(select(func.count(Member.id)))).scalar() or 0
    active_members = (await db.execute(select(func.count(Member.id)).where(Member.is_active == True))).scalar() or 0
    total_causes = (await db.execute(select(func.count(ContributionCause.id)))).scalar() or 0
    total_contributions = (await db.execute(select(func.count(Contribution.id)))).scalar() or 0
    total_collected = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)))).scalar() or 0)

    # Top 10
    top_q = (await db.execute(
        select(Member.id, Member.name, Member.member_number,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.id).label("count"))
        .outerjoin(Contribution, Contribution.member_id == Member.id)
        .group_by(Member.id, Member.name, Member.member_number)
        .order_by(desc("total")).limit(10))).all()
    top_members = [{"id": r.id, "name": r.name, "member_number": r.member_number, "total": float(r.total), "count": r.count} for r in top_q]

    # Per-cause breakdown
    causes_result = await db.execute(
        select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.id.distinct()).label("contributors"))
        .outerjoin(Contribution, Contribution.cause_id == ContributionCause.id)
        .group_by(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount))
    cause_stats = []
    for r in causes_result:
        pct = round(float(r.total) / float(r.target_amount) * 100, 1) if r.target_amount and r.target_amount > 0 else None
        cause_stats.append({"id": r.id, "name": r.name, "total": float(r.total), "contributors": r.contributors, "target": float(r.target_amount) if r.target_amount else 0, "progress": pct})

    # Monthly trends (last 6 months)
    from sqlalchemy import extract
    current_year = 2026
    months = []
    month_labels = []
    month_data = []
    for m in range(1, 8):  # Jan - Jul 2026
        monthly = (await db.execute(
            select(func.coalesce(func.sum(Contribution.amount), 0))
            .where(extract('year', Contribution.date_paid) == current_year)
            .where(extract('month', Contribution.date_paid) == m)
        )).scalar() or 0
        months.append({"month": m, "total": float(monthly)})
        from datetime import date as dt
        month_labels.append(date(2026, m, 1).strftime("%b"))
        month_data.append(float(monthly))

    # Payment method breakdown
    methods = ["cash", "mpesa", "bank"]
    method_data = {}
    for m in methods:
        t = (await db.execute(
            select(func.coalesce(func.sum(Contribution.amount), 0))
            .where(Contribution.payment_method == m)
        )).scalar() or 0
        method_data[m] = float(t)

    # Alerts
    alerts = []
    # Causes near target (80%+)
    for c in cause_stats:
        if c["target"] > 0 and c["progress"] and c["progress"] >= 80:
            alerts.append({"type": "success" if c["progress"] >= 100 else "warning",
                          "icon": "trophy" if c["progress"] >= 100 else "chart-line",
                          "msg": f"<strong>{c['name']}</strong> is {'fully funded' if c['progress'] >= 100 else str(c['progress'])+'% funded'} (KES {c['total']:,.0f} of KES {c['target']:,.0f})"})
    # Causes with balance (not fully disbursed)
    for c in cause_stats:
        if c["total"] > 0:
            disbursed = float((await db.execute(select(func.coalesce(func.sum(Disbursement.amount), 0)).where(Disbursement.cause_id == c["id"]))).scalar() or 0)
            if disbursed == 0 and c["total"] > 0:
                alerts.append({"type": "info", "icon": "money-bill-wave",
                              "msg": f"<strong>{c['name']}</strong> raised KES {c['total']:,.0f} — no disbursements recorded yet"})

    return render("dashboard.html", user=user, request=request,
        stats={"total_members": total_members, "active_members": active_members, "inactive_members": total_members - active_members,
               "total_causes": total_causes, "total_contributions": total_contributions,
               "total_collected": total_collected, "avg_per_member": round(total_collected / active_members, 0) if active_members else 0,
               "avg_per_cause": round(total_collected / total_causes, 0) if total_causes else 0},
        top_members=top_members, cause_stats=cause_stats,
        chart_labels=[c["name"][:25] for c in cause_stats],
        chart_data=[c["total"] for c in cause_stats],
        month_labels=month_labels, month_data=month_data,
        method_data=method_data, alerts=alerts)


# Members list
@app.get("/alumni", response_class=HTMLResponse)
async def member_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    search = request.query_params.get("q", "")
    q = select(Member).where(Member.name.ilike(f"%{search}%")).order_by(Member.member_number) if search else select(Member).order_by(Member.member_number)
    result = await db.execute(q)
    member_data = []
    for m in result.scalars().all():
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == m.id))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.id)).where(Contribution.member_id == m.id))).scalar() or 0
        member_data.append({"id": m.id, "member_number": m.member_number, "name": m.name, "phone_number": m.phone_number, "photo": m.photo, "is_active": m.is_active, "total": total, "count": count})
    return render("members.html", user=user, request=request, members=member_data, search=search)


@app.get("/alumni/register", response_class=HTMLResponse)
async def member_new_form(request: Request, user: str = Depends(require_auth)):
    return render("member_form.html", user=user, request=request, member=None)


@app.post("/alumni/register")
async def member_create(name: str = Form(...), phone_number: str = Form(""), db: AsyncSession = Depends(get_db), user: str = Depends(require_admin)):
    last = (await db.execute(select(func.max(Member.member_number)))).scalar() or 0
    db.add(Member(member_number=last + 1, name=name.strip(), phone_number=phone_number.strip()))
    await db.commit()
    return RedirectResponse(url="/alumni", status_code=302)


@app.get("/alumni/{member_id}", response_class=HTMLResponse)
async def member_detail(member_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    contribs = await db.execute(select(Contribution).where(Contribution.member_id == member_id).options(selectinload(Contribution.cause)).order_by(desc(Contribution.date_paid)))
    total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == member_id))).scalar() or 0)
    causes = (await db.execute(select(ContributionCause).where(ContributionCause.is_active == True))).scalars().all()
    return render("member_detail.html", user=user, request=request, member=member, contributions=contribs.scalars().all(), total=total, causes=causes, today=date.today().isoformat())


@app.post("/alumni/{member_id}/edit")
async def member_update(member_id: int, name: str = Form(...), phone_number: str = Form(""), is_active: bool = Form(False), db: AsyncSession = Depends(get_db), user: str = Depends(require_admin)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    member.name, member.phone_number, member.is_active = name.strip(), phone_number.strip(), is_active
    await db.commit()
    return RedirectResponse(url=f"/alumni/{member_id}", status_code=302)


@app.post("/alumni/{member_id}/add-contribution")
async def member_add_contribution(member_id: int, cause_id: int = Form(...), amount: float = Form(...),
    payment_method: str = Form("cash"), transaction_ref: str = Form(""),
    date_paid: str = Form(""), notes: str = Form(""),
    db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    from datetime import date as date_cls
    try: dp = date_cls.fromisoformat(date_paid)
    except: dp = date_cls.today()
    db.add(Contribution(member_id=member_id, cause_id=cause_id, amount=amount,
        payment_method=payment_method, transaction_ref=transaction_ref, date_paid=dp, notes=notes))
    await db.commit()
    return await member_detail(member_id, request, db, user)


# ── Member photo upload ──
import uuid as _uuid
import aiofiles

UPLOAD_DIR = Path(__file__).parent / "static" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@app.post("/alumni/{member_id}/photo", response_class=HTMLResponse)
async def member_photo_upload(
    member_id: int, file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db), user: str = Depends(require_admin),
):
    member = await db.get(Member, member_id)
    if not member:
        return HTMLResponse('<div class="alert alert-danger">Member not found</div>')

    # Validate file type
    allowed = ("image/jpeg", "image/png", "image/webp", "image/gif")
    if file.content_type not in allowed:
        return HTMLResponse(f'<div class="alert alert-danger">Invalid file type. Allowed: {", ".join(a.split("/")[1] for a in allowed)}</div>')

    # Save file
    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "jpg"
    filename = f"member_{member_id}_{_uuid.uuid4().hex[:8]}.{ext}"
    dest = UPLOAD_DIR / filename

    content = await file.read()
    async with aiofiles.open(str(dest), "wb") as f:
        await f.write(content)

    # Update member photo field
    member.photo = f"/static/uploads/{filename}"
    await db.commit()

    return HTMLResponse(f'''
    <div class="alert alert-success">Photo uploaded successfully</div>
    <img src="/static/uploads/{filename}" class="rounded-circle mt-2" style="width:120px;height:120px;object-fit:cover;border:3px solid var(--accent)">
    ''')



# Causes
@app.get("/welfare-causes", response_class=HTMLResponse)
async def cause_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    result = await db.execute(
        select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount, ContributionCause.is_active,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.member_id.distinct()).label("contributors"))
        .outerjoin(Contribution, Contribution.cause_id == ContributionCause.id)
        .group_by(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount, ContributionCause.is_active).order_by(ContributionCause.id))
    causes = []
    for r in result:
        pct = round(float(r.total) / float(r.target_amount) * 100, 1) if r.target_amount and r.target_amount > 0 else None
        causes.append({"id": r.id, "name": r.name, "total": float(r.total), "target": float(r.target_amount) if r.target_amount else 0, "contributors": r.contributors, "active": r.is_active, "progress": pct})
    return render("causes.html", user=user, request=request, causes=causes)


# Contributions
@app.get("/contribution-records", response_class=HTMLResponse)
async def contribution_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    result = await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(desc(Contribution.date_paid)).limit(200))
    causes = (await db.execute(select(ContributionCause).order_by(ContributionCause.name))).scalars().all()
    return render("contributions.html", user=user, request=request, contributions=result.scalars().all(), causes=causes)


# Export CSV
@app.get("/exports/csv")
async def export_csv(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    import csv, io
    result = await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(Contribution.date_paid))
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Member #", "Member Name", "Cause", "Amount (KES)", "Date Paid", "Notes"])
    for c in result.scalars().all():
        w.writerow([c.member.member_number, c.member.name, c.cause.name, float(c.amount), c.date_paid.isoformat(), c.notes])
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=kh07_contributions.csv"})


# Export Excel
@app.get("/exports/excel")
async def export_excel(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    def sh(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h); cell.font, cell.fill, cell.border = h_font, h_fill, thin
    def aw(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 50)
    
    ws = wb.active; ws.title = "Contributions"
    sh(ws, ["Member #", "Member Name", "Cause", "Amount (KES)", "Date Paid", "Notes"])
    for i, c in enumerate((await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(Contribution.date_paid))).scalars().all(), 2):
        ws.cell(row=i, column=1, value=c.member.member_number).border = thin
        ws.cell(row=i, column=2, value=c.member.name).border = thin
        ws.cell(row=i, column=3, value=c.cause.name).border = thin
        a = ws.cell(row=i, column=4, value=float(c.amount)); a.number_format = '#,##0'; a.border = thin
        ws.cell(row=i, column=5, value=c.date_paid.isoformat()).border = thin
        ws.cell(row=i, column=6, value=c.notes).border = thin
    aw(ws)
    
    ws2 = wb.create_sheet("Member Summary")
    sh(ws2, ["#", "Name", "Phone", "Total (KES)", "Contributions", "Active"])
    for i, m in enumerate((await db.execute(select(Member).order_by(Member.member_number))).scalars().all(), 2):
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == m.id))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.id)).where(Contribution.member_id == m.id))).scalar() or 0
        ws2.cell(row=i, column=1, value=m.member_number).border = thin
        ws2.cell(row=i, column=2, value=m.name).border = thin
        ws2.cell(row=i, column=3, value=m.phone_number).border = thin
        a = ws2.cell(row=i, column=4, value=total); a.number_format = '#,##0'; a.border = thin
        ws2.cell(row=i, column=5, value=count).border = thin
        ws2.cell(row=i, column=6, value="Yes" if m.is_active else "No").border = thin
    aw(ws2)
    
    ws3 = wb.create_sheet("Cause Summary")
    sh(ws3, ["Cause", "Total (KES)", "Contributors", "Target (KES)", "Progress (%)"])
    for i, (cid, cname, tgt) in enumerate((await db.execute(select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount))), 2):
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.cause_id == cid))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.member_id.distinct())).where(Contribution.cause_id == cid))).scalar() or 0
        ws3.cell(row=i, column=1, value=cname).border = thin
        a = ws3.cell(row=i, column=2, value=total); a.number_format = '#,##0'; a.border = thin
        ws3.cell(row=i, column=3, value=count).border = thin
        t = ws3.cell(row=i, column=4, value=float(tgt) if tgt else 0); t.number_format = '#,##0'; t.border = thin
        ws3.cell(row=i, column=5, value=round(float(total) / float(tgt) * 100, 1) if tgt and tgt > 0 else 0).border = thin
    aw(ws3)
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=kh07_contributions.xlsx"})


# ── Filtered contributions (HTMX) ──
@app.get("/contribution-records/filter", response_class=HTMLResponse)
async def contributions_filtered(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause_id = request.query_params.get("cause_id", "")
    month = request.query_params.get("month", "")
    year = request.query_params.get("year", "")
    method = request.query_params.get("method", "")
    q_text = request.query_params.get("q", "").strip()
    amt_min = request.query_params.get("amount_min", "")
    amt_max = request.query_params.get("amount_max", "")
    
    q = select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause))
    
    if cause_id and cause_id.isdigit():
        q = q.where(Contribution.cause_id == int(cause_id))
    if month and month.isdigit():
        from sqlalchemy import extract
        q = q.where(extract("month", Contribution.date_paid) == int(month))
    if year and year.isdigit():
        from sqlalchemy import extract
        q = q.where(extract("year", Contribution.date_paid) == int(year))
    if method:
        q = q.where(Contribution.payment_method == method)
    if q_text:
        pattern = f"%{q_text}%"
        q = q.join(Contribution.member).where(Member.name.ilike(pattern) | Member.member_number.cast(String).ilike(pattern))
    if amt_min and amt_min.replace(".", "").isdigit():
        q = q.where(Contribution.amount >= float(amt_min))
    if amt_max and amt_max.replace(".", "").isdigit():
        q = q.where(Contribution.amount <= float(amt_max))
    
    q = q.order_by(desc(Contribution.date_paid)).limit(200)
    result = await db.execute(q)
    return render("_contrib_table.html", user=user, request=request, contributions=result.scalars().all())


# ── Member statement ──
@app.get("/alumni/{member_id}/statement", response_class=HTMLResponse)
async def member_statement(member_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    
    contribs = await db.execute(
        select(Contribution).where(Contribution.member_id == member_id)
        .options(selectinload(Contribution.cause))
        .order_by(Contribution.date_paid)
    )
    contributions = contribs.scalars().all()
    
    total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == member_id))).scalar() or 0)
    
    # Per-cause totals
    cause_totals = {}
    for c in contributions:
        name = c.cause.name
        cause_totals[name] = cause_totals.get(name, 0) + float(c.amount)
    
    return render("statement.html", user=user, request=request, member=member,
                  contributions=contributions, total=total, cause_totals=cause_totals)


# ── Inline edit member name (HTMX) ──
@app.post("/alumni/{member_id}/inline-edit")
async def member_inline_edit(member_id: int, request: Request, field: str = Form(...), value: str = Form(""), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    
    if field == "name":
        member.name = value.strip()
    elif field == "phone":
        member.phone_number = value.strip()
    elif field == "is_active":
        member.is_active = value.lower() in ("true", "1", "yes")
    
    await db.commit()
    return HTMLResponse(value.strip() if value.strip() else "—")


# ── Import Excel (upload form) ──
@app.get("/bulk-upload", response_class=HTMLResponse)
async def import_form(request: Request, user: str = Depends(require_auth)):
    return render("import.html", user=user, request=request)


@app.post("/bulk-upload", response_class=HTMLResponse)
async def import_excel(request: Request, file: UploadFile = File(...), user: str = Depends(require_admin)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        return HTMLResponse('<div class="alert alert-danger">Please upload a .xlsx file</div>')

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Detect format: headers in row 1
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    async with async_session() as db:
        # Create missing causes from headers (col 3+)
        cause_map = {}
        for i, h in enumerate(headers[2:], 3):
            if h and str(h).strip():
                name = str(h).strip().split("(")[0].strip()
                existing = (await db.execute(select(ContributionCause).where(ContributionCause.name == name))).scalar_one_or_none()
                if not existing:
                    existing = ContributionCause(name=name)
                    db.add(existing)
                    await db.flush()
                cause_map[i] = existing.id

        imported = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            name = str(row[0]).strip()
            phone = str(row[1]).strip() if row[1] else ""

            member = (await db.execute(select(Member).where(Member.name == name))).scalar_one_or_none()
            if not member:
                member = Member(name=name, phone_number=phone)
                db.add(member)
                await db.flush()

            for col_idx, cause_id in cause_map.items():
                amount = row[col_idx - 1]
                if amount and float(amount) > 0:
                    existing_contrib = (await db.execute(
                        select(Contribution).where(
                            Contribution.member_id == member.id,
                            Contribution.cause_id == cause_id,
                        ).limit(1)
                    )).scalar_one_or_none()
                    if not existing_contrib:
                        db.add(Contribution(
                            member_id=member.id, cause_id=cause_id,
                            amount=float(amount), date_paid=date.today(),
                            payment_method="cash",
                        ))
                        imported += 1
                    else:
                        existing_contrib.amount = float(amount)

        await db.commit()
        total_members = (await db.execute(select(func.count(Member.id)))).scalar()

    return HTMLResponse(f"""<div class="alert alert-success">
        <i class="fas fa-check-circle me-2"></i>Import complete!
        <br><strong>{imported}</strong> contributions imported
        <br><strong>{total_members}</strong> total members
        <br class="small text-muted">{len(errors)} errors
    </div>""")


# ── Cause edit (target amount) ──
@app.post("/welfare-causes/{cause_id}/edit")
async def cause_edit(cause_id: int, name: str = Form(...), target_amount: float = Form(0), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = await db.get(ContributionCause, cause_id)
    if not cause:
        raise HTTPException(status_code=404)
    cause.name = name.strip()
    cause.target_amount = target_amount
    await db.commit()
    return RedirectResponse(url="/welfare-causes", status_code=302)


# ── Contribution receipt ──
@app.get("/payment-receipt/{contrib_id}", response_class=HTMLResponse)
async def contribution_receipt(contrib_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    c = await db.get(Contribution, contrib_id)
    if not c:
        raise HTTPException(status_code=404)
    member = await db.get(Member, c.member_id)
    cause = await db.get(ContributionCause, c.cause_id)
    return render("receipt.html", user=user, request=request, c=c, member=member, cause=cause)


# ── Contribution edit ──
@app.get("/contribution-records/{contrib_id}/edit", response_class=HTMLResponse)
async def contribution_edit_form(contrib_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    c = await db.get(Contribution, contrib_id)
    if not c: raise HTTPException(status_code=404)
    causes = (await db.execute(select(ContributionCause).order_by(ContributionCause.name))).scalars().all()
    return render("contrib_edit.html", user=user, request=request, c=c, causes=causes,
                  today=c.date_paid.isoformat())

@app.post("/contribution-records/{contrib_id}/edit")
async def contribution_edit(contrib_id: int, cause_id: int = Form(...), amount: float = Form(...),
    payment_method: str = Form("cash"), transaction_ref: str = Form(""),
    date_paid: str = Form(""), notes: str = Form(""),
    db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    c = await db.get(Contribution, contrib_id)
    if not c: raise HTTPException(status_code=404)
    from datetime import date as dc
    c.cause_id = cause_id; c.amount = amount; c.payment_method = payment_method
    c.transaction_ref = transaction_ref; c.notes = notes
    try: c.date_paid = dc.fromisoformat(date_paid)
    except: pass
    await db.commit()
    return RedirectResponse(url=f"/members/{c.member_id}", status_code=302)


# ── Contribution delete ──
@app.post("/contribution-records/{contrib_id}/delete")
async def contribution_delete(contrib_id: int, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    c = await db.get(Contribution, contrib_id)
    if not c: raise HTTPException(status_code=404)
    mid = c.member_id
    await db.delete(c)
    await db.commit()
    return RedirectResponse(url=f"/members/{mid}", status_code=302)


# ── Filtered export (HTMX partial for export) ──
@app.get("/exports/filtered-csv")
async def export_filtered_csv(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause_id = request.query_params.get("cause_id", "")
    month = request.query_params.get("month", "")
    q = select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause))
    if cause_id and cause_id.isdigit(): q = q.where(Contribution.cause_id == int(cause_id))
    if month and month.isdigit():
        from sqlalchemy import extract
        q = q.where(extract("month", Contribution.date_paid) == int(month))
    contributions = (await db.execute(q.order_by(Contribution.date_paid))).scalars().all()
    
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Member", "Cause", "Amount (KES)", "Method", "Ref", "Notes"])
    for c in contributions:
        w.writerow([c.date_paid.isoformat(), c.member.name, c.cause.name, float(c.amount), c.payment_method, c.transaction_ref, c.notes])
    return Response(content=buf.getvalue(), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=kh07_filtered_contributions.csv"})


# ── Dashboard stats filtered (HTMX partial) ──
@app.get("/overview/stats", response_class=HTMLResponse)
async def dashboard_stats_partial(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    from_date = request.query_params.get("from", "")
    to_date = request.query_params.get("to", "")
    from datetime import date as dc
    
    try: fd = dc.fromisoformat(from_date) if from_date else None
    except: fd = None
    try: td = dc.fromisoformat(to_date) if to_date else None
    except: td = None
    
    base_q = select(func.coalesce(func.sum(Contribution.amount), 0))
    if fd: base_q = base_q.where(Contribution.date_paid >= fd)
    if td: base_q = base_q.where(Contribution.date_paid <= td)
    
    total_collected = float((await db.execute(base_q)).scalar() or 0)
    
    total_members_q = select(func.count(Member.id))
    total_members = (await db.execute(total_members_q)).scalar() or 0
    
    cnt_q = select(func.count(Contribution.id))
    if fd: cnt_q = cnt_q.where(Contribution.date_paid >= fd)
    if td: cnt_q = cnt_q.where(Contribution.date_paid <= td)
    total_contributions = (await db.execute(cnt_q)).scalar() or 0
    
    return HTMLResponse(f"""<div class="row g-3 mb-4" id="dashboard-stats">
        <div class="col-xl-3 col-md-6"><div class="stat-card accent">
            <i class="fas fa-users icon"></i>
            <div class="value">{total_members}</div>
            <div class="label">Total Members</div>
        </div></div>
        <div class="col-xl-3 col-md-6"><div class="stat-card success">
            <i class="fas fa-coins icon"></i>
            <div class="value">KES {"{:,.0f}".format(total_collected)}</div>
            <div class="label">Total Collected</div>
        </div></div>
        <div class="col-xl-3 col-md-6"><div class="stat-card info">
            <i class="fas fa-hand-holding-heart icon"></i>
            <div class="value">{total_contributions}</div>
            <div class="label">Contributions</div>
        </div></div>
        <div class="col-xl-3 col-md-6"><div class="stat-card warning">
            <i class="fas fa-chart-line icon"></i>
            <div class="value">KES {"{:,.0f}".format(round(total_collected / total_members, 0)) if total_members else 0}</div>
            <div class="label">Avg per Member</div>
        </div></div>
    </div>""")


# ── Disbursement routes ──
@app.get("/welfare-causes/{cause_id}/disburse", response_class=HTMLResponse)
async def disburse_form(cause_id: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = await db.get(ContributionCause, cause_id)
    if not cause: raise HTTPException(status_code=404)
    total_raised = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.cause_id == cause_id))).scalar() or 0)
    total_disbursed = float((await db.execute(select(func.coalesce(func.sum(Disbursement.amount), 0)).where(Disbursement.cause_id == cause_id))).scalar() or 0)
    disbursements = (await db.execute(select(Disbursement).where(Disbursement.cause_id == cause_id).order_by(desc(Disbursement.date_disbursed)))).scalars().all()
    today_str = date.today().isoformat()
    return render("disburse.html", user=user, request=request, cause=cause, total_raised=total_raised,
                  total_disbursed=total_disbursed, balance=total_raised - total_disbursed, disbursements=disbursements, today=today_str)


@app.post("/welfare-causes/{cause_id}/disburse")
async def disburse_create(cause_id: int, beneficiary_name: str = Form(...), amount: float = Form(...),
    date_disbursed: str = Form(""), notes: str = Form(""), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = await db.get(ContributionCause, cause_id)
    if not cause: raise HTTPException(status_code=404)
    from datetime import date as dc
    try: dd = dc.fromisoformat(date_disbursed)
    except: dd = dc.today()
    db.add(Disbursement(cause_id=cause_id, beneficiary_name=beneficiary_name.strip(), amount=amount, date_disbursed=dd, notes=notes))
    await db.commit()
    return RedirectResponse(url=f"/welfare-causes/{cause_id}/disburse", status_code=302)


# ── Telegram notification (triggered when cause created) ──
def send_telegram(message: str):
    import urllib.request, urllib.parse
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "7605394619:***")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "6760963523")
    try:
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": message, "parse_mode": "HTML"}).encode()
        urllib.request.urlopen(f"https://api.telegram.org/bot{token}/sendMessage", data=data, timeout=10)
    except:
        pass  # silently fail


@app.post("/welfare-causes/new")
async def cause_create(name: str = Form(...), target_amount: float = Form(0), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = ContributionCause(name=name.strip(), target_amount=target_amount if target_amount > 0 else None)
    db.add(cause)
    await db.commit()
    AuditLog.add("CREATE CAUSE", f"{name.strip()} (KES {target_amount:,.0f})")
    # Send Telegram notification
    msg = f"""<b>🆕 New Welfare Cause</b>
<b>{name.strip()}</b>
Target: KES {target_amount:,.0f}
<a href="https://kh07-welfare.spidmax.win">View Details</a>"""
    send_telegram(msg)
    return RedirectResponse(url="/welfare-causes", status_code=302)


# ── Annual report ──
@app.get("/annual-report/{year}", response_class=HTMLResponse)
async def annual_report(year: int, request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    from sqlalchemy import extract
    collected = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0))
        .where(extract("year", Contribution.date_paid) == year))).scalar() or 0)
    contrib_count = (await db.execute(select(func.count(Contribution.id))
        .where(extract("year", Contribution.date_paid) == year))).scalar() or 0
    member_count = (await db.execute(select(func.count(Member.id)))).scalar() or 0
    active_count = (await db.execute(select(func.count(Member.id)).where(Member.is_active == True))).scalar() or 0
    
    # Per-cause breakdown for the year
    causes_data = (await db.execute(
        select(ContributionCause.id, ContributionCause.name, func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.id).label("count"))
        .outerjoin(Contribution, (Contribution.cause_id == ContributionCause.id) & (extract("year", Contribution.date_paid) == year))
        .group_by(ContributionCause.id, ContributionCause.name))).all()
    cause_stats = [{"name": r.name, "total": float(r.total), "count": r.count} for r in causes_data if r.total > 0]
    
    # Top contributors
    top = (await db.execute(
        select(Member.name, func.coalesce(func.sum(Contribution.amount), 0).label("total"))
        .outerjoin(Contribution, (Contribution.member_id == Member.id) & (extract("year", Contribution.date_paid) == year))
        .group_by(Member.id, Member.name).order_by(desc("total")).limit(10))).all()
    top_members = [{"name": r.name, "total": float(r.total)} for r in top if r.total > 0]
    
    # Disbursements
    total_disbursed = float((await db.execute(select(func.coalesce(func.sum(Disbursement.amount), 0))
        .where(extract("year", Disbursement.date_disbursed) == year))).scalar() or 0)
    
    monthly = []
    for m in range(1, 13):
        t = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0))
            .where(extract("year", Contribution.date_paid) == year).where(extract("month", Contribution.date_paid) == m))).scalar() or 0)
        monthly.append(t)
    
    from datetime import date as dc
    month_labels = [dc(year, m, 1).strftime("%b") for m in range(1, 13)]
    
    return render("report.html", user=user, request=request, year=year, collected=collected,
                  contrib_count=contrib_count, member_count=member_count, active_count=active_count,
                  cause_stats=cause_stats, top_members=top_members, total_disbursed=total_disbursed,
                  monthly=monthly, month_labels=month_labels)


# ── PDF receipt download ──
@app.get("/payment-receipt/{contrib_id}/pdf")
async def receipt_pdf(contrib_id: int, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    c = await db.get(Contribution, contrib_id)
    if not c: raise HTTPException(status_code=404)
    member = await db.get(Member, c.member_id)
    cause = await db.get(ContributionCause, c.cause_id)
    
    from fpdf import FPDF
    pdf = FPDF(orientation="P", unit="mm", format=(80, 120))
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)
    
    pdf.set_fill_color(184, 67, 58)
    pdf.rect(0, 0, 80, 12, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_xy(0, 3)
    pdf.cell(80, 6, "KH07 ALUMNI WELFARE", align="C", ln=True)
    
    pdf.set_text_color(40, 40, 40)
    pdf.set_y(16)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(80, 5, "CONTRIBUTION RECEIPT", align="C", ln=True)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(80, 4, f"Receipt #{c.id}  |  {c.date_paid}", align="C", ln=True)
    
    pdf.set_y(28)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(20, 5, "Member:")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(55, 5, member.name, ln=True)
    
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(20, 5, "Cause:")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(55, 5, cause.name[:40], ln=True)
    
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(20, 5, "Amount:")
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(184, 67, 58)
    pdf.cell(55, 7, f"KES {float(c.amount):,.0f}", ln=True)
    
    pdf.set_text_color(40, 40, 40)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(20, 5, "Method:")
    pdf.set_font("Helvetica", "", 8)
    method = c.payment_method.upper() + (f" ({c.transaction_ref})" if c.transaction_ref else "")
    pdf.cell(55, 5, method, ln=True)
    
    if c.notes:
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(20, 5, "Notes:")
        pdf.set_font("Helvetica", "", 7)
        pdf.multi_cell(55, 4, c.notes, ln=True)
    
    pdf.set_y(95)
    pdf.set_draw_color(184, 67, 58)
    pdf.line(5, pdf.get_y(), 75, pdf.get_y())
    pdf.set_y(98)
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(70, 4, "Thank you for your contribution.", align="C", ln=True)
    pdf.cell(70, 3, "This is a computer-generated receipt.", align="C")
    
    return Response(content=bytes(pdf.output("", dest="S")),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=receipt_{contrib_id}.pdf"})


# ── Cause archive (toggle active) ──
@app.post("/welfare-causes/{cause_id}/archive")
async def cause_archive(cause_id: int, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    cause = await db.get(ContributionCause, cause_id)
    if not cause: raise HTTPException(status_code=404)
    cause.is_active = not cause.is_active
    await db.commit()
    return RedirectResponse(url="/welfare-causes", status_code=302)


# ── Database backup download ──
@app.get("/data-backup")
async def db_backup(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    from app.database import DATABASE_URL
    db_path = DATABASE_URL.replace("sqlite+aiosqlite:///", "")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
    size = os.path.getsize(db_path)
    return Response(content=open(db_path, "rb").read(),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename=kh07_welfare_backup_{date.today().isoformat()}.sqlite3",
                 "Content-Length": str(size)})


# ── Audit log model ──
class AuditLog:
    _entries: list = []
    MAX = 500

    @classmethod
    def add(cls, action: str, details: str, username: str = "admin"):
        cls._entries.append({"time": datetime.now().isoformat(), "action": action, "details": details, "user": username})
        if len(cls._entries) > cls.MAX:
            cls._entries = cls._entries[-cls.MAX:]

    @classmethod
    def recent(cls, limit: int = 50):
        return list(reversed(cls._entries[-limit:]))


@app.get("/activity-log", response_class=HTMLResponse)
async def audit_page(request: Request, user: str = Depends(require_auth)):
    return render("audit.html", user=user, request=request, entries=AuditLog.recent(100))


# ── Admin: User Management ──
@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    result = await db.execute(select(User).order_by(User.username))
    users = result.scalars().all()
    return render("admin_users.html", request=request, user=admin.username, users=users)


@app.post("/admin/users/create", response_class=HTMLResponse)
async def admin_create_user(
    username: str = Form(...), password: str = Form(...), role: str = Form("viewer"),
    db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin),
):
    from app.auth import hash_password
    existing = (await db.execute(select(User).where(User.username == username))).scalar_one_or_none()
    if existing:
        return HTMLResponse(f'<div class="alert alert-danger">User "{username}" already exists</div>')
    user = User(username=username, password_hash=hash_password(password), role=role)
    db.add(user)
    await db.commit()
    return HTMLResponse(f'<div class="alert alert-success">Created user <strong>{username}</strong> ({role})</div>')


@app.post("/admin/users/{user_id}/toggle", response_class=HTMLResponse)
async def admin_toggle_user(
    user_id: int, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        return HTMLResponse('<div class="alert alert-danger">User not found</div>')
    if user.username == "admin":
        return HTMLResponse('<div class="alert alert-warning">Cannot deactivate the main admin</div>')
    user.is_active = not user.is_active
    await db.commit()
    status = "activated" if user.is_active else "deactivated"
    return HTMLResponse(f'<div class="alert alert-success">User <strong>{user.username}</strong> {status}</div>')


@app.post("/admin/users/{user_id}/delete", response_class=HTMLResponse)
async def admin_delete_user(
    user_id: int, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        return HTMLResponse('<div class="alert alert-danger">User not found</div>')
    if user.username == "admin":
        return HTMLResponse('<div class="alert alert-warning">Cannot delete the main admin</div>')
    await db.delete(user)
    await db.commit()
    return HTMLResponse(f'<div class="alert alert-success">User <strong>{user.username}</strong> deleted</div>')


@app.post("/admin/users/{user_id}/reset-password", response_class=HTMLResponse)
async def admin_reset_password(
    user_id: int, new_password: str = Form(...), db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
):
    from app.auth import hash_password
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        return HTMLResponse('<div class="alert alert-danger">User not found</div>')
    user.password_hash = hash_password(new_password)
    await db.commit()
    return HTMLResponse(f'<div class="alert alert-success">Password reset for <strong>{user.username}</strong></div>')


# ── M-Pesa Integration ──

@app.get("/admin/mpesa", response_class=HTMLResponse)
async def mpesa_admin_page(request: Request, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    """M-Pesa configuration page."""
    cfg = await db.get(MpesaConfig, 1)
    return render("admin_mpesa.html", request=request, user=admin.username, cfg=cfg)


@app.post("/admin/mpesa/save", response_class=HTMLResponse)
async def mpesa_admin_save(
    consumer_key: str = Form(""), consumer_secret: str = Form(""),
    passkey: str = Form(""), shortcode: str = Form("174379"),
    callback_url: str = Form(""), sandbox: bool = Form(True),
    db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin),
):
    cfg = await db.get(MpesaConfig, 1)
    if not cfg:
        cfg = MpesaConfig(id=1)
        db.add(cfg)
    if consumer_key:
        cfg.consumer_key = consumer_key
    if consumer_secret:
        cfg.consumer_secret = consumer_secret
    if passkey:
        cfg.passkey = passkey
    cfg.shortcode = shortcode or "174379"
    cfg.callback_url = callback_url
    cfg.sandbox = sandbox
    await db.commit()
    return HTMLResponse('<div class="alert alert-success"><i class="fas fa-check-circle me-1"></i>M-Pesa settings saved</div>')


@app.get("/mpesa/transactions", response_class=HTMLResponse)
async def mpesa_transactions(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    """M-Pesa transaction history page."""
    page = int(request.query_params.get("page", "1"))
    per_page = 50
    offset = (page - 1) * per_page
    total = (await db.execute(select(func.count(MpesaTransaction.id)))).scalar() or 0
    result = await db.execute(
        select(MpesaTransaction)
        .options(selectinload(MpesaTransaction.member), selectinload(MpesaTransaction.cause))
        .order_by(desc(MpesaTransaction.created_at))
        .offset(offset).limit(per_page)
    )
    txs = result.scalars().all()
    return render("mpesa_transactions.html", request=request, user=user,
                  transactions=txs, page=page, total=total, per_page=per_page)


@app.post("/mpesa/stkpush", response_class=HTMLResponse)
async def mpesa_stk_push(
    member_id: int = Form(...), cause_id: int = Form(...), amount: float = Form(...),
    phone: str = Form(...), db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin),
):
    """Initiate an M-Pesa STK Push payment request."""
    from app.mpesa import stk_push, log_transaction, _format_phone

    member = await db.get(Member, member_id)
    if not member:
        return HTMLResponse('<div class="alert alert-danger">Member not found</div>')

    try:
        phone = _format_phone(phone)
    except Exception as e:
        return HTMLResponse(f'<div class="alert alert-danger">{str(e)}</div>')

    try:
        account_ref = f"KH{member.member_number}"
        result = await stk_push(phone=phone, amount=amount, account_ref=account_ref)

        code = result.get("ResponseCode", "1")
        if code == "0":
            checkout_id = result.get("CheckoutRequestID", "")
            merchant_id = result.get("MerchantRequestID", "")

            # Log transaction
            await log_transaction(db, checkout_id, merchant_id, member_id, cause_id, amount, phone, account_ref)

            return HTMLResponse(f"""
            <div class="alert alert-success">
                <i class="fas fa-mobile-alt me-1"></i>STK Push sent to <strong>{phone}</strong>!<br>
                <small>Check your phone to complete payment of <strong>KES {amount:,.0f}</strong></small><br>
                <small class="text-muted">Ref: {checkout_id[:15]}…</small>
            </div>
            <div class="mt-2" id="mpesa-poll-{checkout_id}">
                <button class="btn btn-sm btn-outline-accent" 
                    hx-get="/mpesa/check/{checkout_id}" 
                    hx-target="#mpesa-poll-{checkout_id}" 
                    hx-swap="outerHTML">
                    <i class="fas fa-sync me-1"></i>Check Payment Status
                </button>
            </div>
            <script>
                setTimeout(function() {{
                    var btn = document.querySelector('[hx-get*="{checkout_id}"]');
                    if (btn) htmx.trigger(btn, 'click');
                }}, 30000);
            </script>
            """)
        else:
            msg = result.get("ResponseDescription", "Unknown error")
            return HTMLResponse(f'<div class="alert alert-danger">M-Pesa request failed: {msg}</div>')
    except Exception as e:
        return HTMLResponse(f'<div class="alert alert-danger">M-Pesa error: {str(e)}</div>')


@app.get("/mpesa/check/{checkout_id}", response_class=HTMLResponse)
async def mpesa_check_status(checkout_id: str, request: Request,
                              db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    """Check the status of an M-Pesa transaction."""
    from app.mpesa import query_status

    try:
        result = await query_status(checkout_id)
        rc = result.get("ResultCode", "1")

        # Update DB
        from app.mpesa import update_transaction
        if rc == "0":
            receipt = result.get("Receipt", "") or next(
                (i.get("Value", "") for i in result.get("CallbackMetadata", {}).get("Item", [])
                 if i.get("Name") == "MpesaReceiptNumber"), "")
            await update_transaction(db, checkout_id, status="success", result_code="0",
                                     result_desc="Completed", receipt=receipt)
        elif rc == "1037":
            await update_transaction(db, checkout_id, status="pending", result_code=rc,
                                     result_desc="Still processing")

        if rc == "0":
            return HTMLResponse(f"""
            <div class="alert alert-success mt-2">
                <i class="fas fa-check-circle me-1"></i>Payment confirmed! KES {result.get("Amount", "?")}<br>
                <small>Receipt: {receipt or '—'}</small>
            </div>""")
        elif rc == "1037":
            return HTMLResponse(f"""
            <div class="alert alert-warning mt-2">
                <i class="fas fa-clock me-1"></i>Still processing. 
                <button class="btn btn-sm btn-outline-accent ms-2"
                    hx-get="/mpesa/check/{checkout_id}" 
                    hx-target="closest div" hx-swap="outerHTML">
                    <i class="fas fa-sync me-1"></i>Check Again
                </button>
            </div>""")
        else:
            await update_transaction(db, checkout_id, status="failed", result_code=rc,
                                     result_desc=result.get("ResultDesc", "Failed"))
            desc = result.get("ResultDesc", "Transaction failed")
            return HTMLResponse(f'<div class="alert alert-danger mt-2">{desc}</div>')
    except Exception as e:
        return HTMLResponse(f'<div class="alert alert-danger mt-2">Check failed: {str(e)}</div>')


@app.post("/api/mpesa/callback")
async def mpesa_callback(request: Request):
    """M-Pesa API callback — receives payment confirmation and auto-reconciles."""
    from app.mpesa import update_transaction, reconcile_from_callback

    try:
        body = await request.json()
        logger.info(f"M-Pesa callback received: {json.dumps(body)[:300]}")

        stk = body.get("Body", {}).get("stkCallback", {})
        checkout_id = stk.get("CheckoutRequestID", "")
        result_code = stk.get("ResultCode", 1)
        result_desc = stk.get("ResultDesc", "")

        async with async_session() as session:
            if result_code == 0:
                meta = stk.get("CallbackMetadata", {}).get("Item", [])
                amount = next((i.get("Value", 0) for i in meta if i.get("Name") == "Amount"), 0)
                receipt = next((i.get("Value", "") for i in meta if i.get("Name") == "MpesaReceiptNumber"), "")
                phone = next((i.get("Value", "") for i in meta if i.get("Name") == "PhoneNumber"), "")

                await reconcile_from_callback(session, checkout_id, amount, receipt, phone)
            else:
                await update_transaction(session, checkout_id, status="failed",
                                         result_code=str(result_code), result_desc=result_desc)

        return {"ResultCode": 0, "ResultDesc": "Accepted"}
    except Exception as e:
        logger.error(f"M-Pesa callback error: {e}")
        return {"ResultCode": 1, "ResultDesc": str(e)}


# ── Viewer role guard middleware ──
@app.middleware("http")
async def viewer_write_guard(request: Request, call_next):
    """Block write operations (POST/PUT/DELETE) for viewer-role users."""
    if request.method in ("POST", "PUT", "DELETE"):
        skip_paths = ("/login", "/self-service/lookup", "/self-service/update-phone", "/self-service/suggest-cause")
        if not request.url.path.startswith(skip_paths):
            from app.auth import get_session_role
            role = get_session_role(request)
            if role == "viewer":
                return HTMLResponse(
                    '<div class="alert alert-danger"><i class="fas fa-ban me-2"></i>'
                    'Viewer accounts cannot perform write operations. Contact an admin.</div>',
                    status_code=403,
                )
    response = await call_next(request)
    return response


# ── Old route redirects ──
_OLD_ROUTES = {
    "/dashboard": "/overview", "/dashboard/stats": "/overview/stats",
    "/import": "/bulk-upload",
    "/members": "/alumni", "/members/new": "/alumni/register",
    "/causes": "/welfare-causes", "/causes/new": "/welfare-causes/new",
    "/contributions": "/contribution-records",
    "/admin/audit": "/activity-log", "/admin/backup": "/data-backup",
    "/portal": "/self-service", "/portal/lookup": "/self-service/lookup",
    "/portal/update-phone": "/self-service/update-phone",
    "/portal/suggest-cause": "/self-service/suggest-cause",
}


@app.middleware("http")
async def redirect_old_routes(request: Request, call_next):
    path = request.url.path
    # Strip trailing slash for matching
    clean = path.rstrip("/")
    if clean in _OLD_ROUTES:
        new_path = _OLD_ROUTES[clean]
        q = request.url.query
        url = new_path + (f"?{q}" if q else "")
        return RedirectResponse(url=url, status_code=301)
    # Handle /members/{id} -> /alumni/{id}, /causes/{id} -> /welfare-causes/{id}
    if clean.startswith("/members/"):
        suffix = clean[9:]
        return RedirectResponse(url=f"/alumni/{suffix}", status_code=301)
    if clean.startswith("/causes/") and not clean.startswith("/welfare-causes/"):
        suffix = clean[8:]
        return RedirectResponse(url=f"/welfare-causes/{suffix}", status_code=301)
    if clean.startswith("/receipt/"):
        suffix = clean[9:]
        return RedirectResponse(url=f"/payment-receipt/{suffix}", status_code=301)
    if clean.startswith("/export/") and not clean.startswith("/exports/"):
        suffix = clean[8:]
        return RedirectResponse(url=f"/exports/{suffix}", status_code=301)
    if clean.startswith("/report/"):
        suffix = clean[8:]
        return RedirectResponse(url=f"/annual-report/{suffix}", status_code=301)
    if clean.startswith("/portal/"):
        suffix = clean[8:]
        return RedirectResponse(url=f"/self-service/{suffix}", status_code=301)
    response = await call_next(request)
    return response


# Exception handlers
@app.exception_handler(404)
async def not_found(request: Request, exc):
    return render("base.html", request=request)


@app.exception_handler(500)
async def server_error(request: Request, exc):
    import traceback
    traceback.print_exc()
    return RedirectResponse(url="/alumni")
