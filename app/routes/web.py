"""Web page routes - rendered with Jinja2 + HTMX."""
from datetime import date, datetime
from decimal import Decimal
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from sqlalchemy import select, func, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import Member, ContributionCause, Contribution
from app.auth import require_auth, verify_password, create_session, logout_session, SESSION_COOKIE, SESSION_MAX_AGE, get_session_user

router = APIRouter()


from jinja2 import Environment, FileSystemLoader
import os

templates_dir = os.path.join(os.path.dirname(__file__), "..", "templates")
env = Environment(loader=FileSystemLoader(templates_dir))


def render(name: str, **ctx) -> HTMLResponse:
    template = env.get_template(name)
    html = template.render(**ctx)
    return HTMLResponse(html)


# Debug - no auth required
@router.get("/ping")
async def ping():
    return {"status": "ok", "time": datetime.now().isoformat()}


@router.get("/check-auth")
async def check_auth(request: Request):
    user = get_session_user(request)
    return {
        "authenticated": user is not None,
        "user": user,
        "cookie_present": SESSION_COOKIE in request.cookies,
    }


# ── Auth ──
@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = ""):
    user = get_session_user(request)
    if user:
        return RedirectResponse(url="/", status_code=302)
    return render("login.html", error=error)


@router.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    if username == "admin" and verify_password(password):
        token = create_session(username)
        resp = RedirectResponse(url="/", status_code=302)
        resp.set_cookie(
            key=SESSION_COOKIE,
            value=token,
            max_age=int(SESSION_MAX_AGE.total_seconds()),
            httponly=True,
            samesite="lax",
        )
        return resp
    return render("login.html", error="Invalid username or password")


@router.get("/logout")
async def logout(request: Request):
    resp = RedirectResponse(url="/login", status_code=302)
    resp.delete_cookie(SESSION_COOKIE)
    return resp


# ── Dashboard ──
@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    # Stats
    total_members = (await db.execute(select(func.count(Member.id)))).scalar() or 0
    active_members = (await db.execute(select(func.count(Member.id)).where(Member.is_active == True))).scalar() or 0
    total_causes = (await db.execute(select(func.count(ContributionCause.id)))).scalar() or 0
    total_contributions = (await db.execute(select(func.count(Contribution.id)))).scalar() or 0
    total_collected = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)))).scalar() or 0)
    
    avg_per_member = round(total_collected / active_members, 0) if active_members else 0
    avg_per_cause = round(total_collected / total_causes, 0) if total_causes else 0

    # Top 10
    top_q = (await db.execute(
        select(Member.id, Member.name, Member.member_number,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.id).label("count"))
        .outerjoin(Contribution, Contribution.member_id == Member.id)
        .group_by(Member.id, Member.name, Member.member_number)
        .order_by(desc("total")).limit(10))).all()
    top_members = [{"id": r.id, "name": r.name, "member_number": r.member_number, "total": float(r.total), "count": r.count} for r in top_q]

    # Per-cause breakdown
    causes_result = await db.execute(
        select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.id.distinct()).label("contributors"))
        .outerjoin(Contribution, Contribution.cause_id == ContributionCause.id)
        .group_by(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount))
    cause_stats = []
    for r in causes_result:
        pct = round(float(r.total) / float(r.target_amount) * 100, 1) if r.target_amount and r.target_amount > 0 else None
        cause_stats.append({"id": r.id, "name": r.name, "total": float(r.total), "contributors": r.contributors, "target": float(r.target_amount) if r.target_amount else 0, "progress": pct})

    return render("dashboard.html", user=user,
        stats={"total_members": total_members, "active_members": active_members, "inactive_members": total_members - active_members,
               "total_causes": total_causes, "total_contributions": total_contributions,
               "total_collected": total_collected, "avg_per_member": avg_per_member, "avg_per_cause": avg_per_cause},
        top_members=top_members, cause_stats=cause_stats,
        chart_labels=[c["name"][:25] for c in cause_stats],
        chart_data=[c["total"] for c in cause_stats])


# ── Members ──
@router.get("/members", response_class=HTMLResponse)
async def member_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    search = request.query_params.get("q", "")
    q = select(Member).where(Member.name.ilike(f"%{search}%")).order_by(Member.member_number) if search else select(Member).order_by(Member.member_number)
    result = await db.execute(q)
    
    member_data = []
    for m in result.scalars().all():
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == m.id))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.id)).where(Contribution.member_id == m.id))).scalar() or 0
        member_data.append({"id": m.id, "member_number": m.member_number, "name": m.name, "phone": m.phone_number, "is_active": m.is_active, "total": total, "count": count})
    
    return render("members.html", user=user, members=member_data, search=search)


@router.get("/members/new", response_class=HTMLResponse)
async def member_new_form(request: Request, user: str = Depends(require_auth)):
    return render("member_form.html", user=user, member=None)


@router.post("/members/new")
async def member_create(name: str = Form(...), phone_number: str = Form(""), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    last = (await db.execute(select(func.max(Member.member_number)))).scalar() or 0
    db.add(Member(member_number=last + 1, name=name.strip(), phone_number=phone_number.strip()))
    await db.commit()
    return RedirectResponse(url="/members", status_code=302)


@router.get("/members/{member_id}", response_class=HTMLResponse)
async def member_detail(member_id: int, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    contribs = await db.execute(select(Contribution).where(Contribution.member_id == member_id).options(selectinload(Contribution.cause)).order_by(desc(Contribution.date_paid)))
    total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == member_id))).scalar() or 0)
    causes = (await db.execute(select(ContributionCause).where(ContributionCause.is_active == True))).scalars().all()
    return render("member_detail.html", user=user, member=member, contributions=contribs.scalars().all(), total=total, causes=causes, today=date.today().isoformat())


@router.post("/members/{member_id}/edit")
async def member_update(member_id: int, name: str = Form(...), phone_number: str = Form(""), is_active: bool = Form(False), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    member = await db.get(Member, member_id)
    if not member:
        raise HTTPException(status_code=404)
    member.name, member.phone_number, member.is_active = name.strip(), phone_number.strip(), is_active
    await db.commit()
    return RedirectResponse(url=f"/members/{member_id}", status_code=302)


@router.post("/members/{member_id}/add-contribution")
async def member_add_contribution(member_id: int, cause_id: int = Form(...), amount: Decimal = Form(...), date_paid: str = Form(...), notes: str = Form(""), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    try:
        dp = date.fromisoformat(date_paid)
    except ValueError:
        dp = date.today()
    db.add(Contribution(member_id=member_id, cause_id=cause_id, amount=amount, date_paid=dp, notes=notes))
    await db.commit()
    return await member_detail(member_id, db, user)


# ── Causes ──
@router.get("/causes", response_class=HTMLResponse)
async def cause_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    result = await db.execute(
        select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount, ContributionCause.is_active,
               func.coalesce(func.sum(Contribution.amount), 0).label("total"),
               func.count(Contribution.member_id.distinct()).label("contributors"))
        .outerjoin(Contribution, Contribution.cause_id == ContributionCause.id)
        .group_by(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount, ContributionCause.is_active).order_by(ContributionCause.id))
    causes = []
    for r in result:
        pct = round(float(r.total) / float(r.target_amount) * 100, 1) if r.target_amount and r.target_amount > 0 else None
        causes.append({"id": r.id, "name": r.name, "total": float(r.total), "target": float(r.target_amount) if r.target_amount else 0, "contributors": r.contributors, "active": r.is_active, "progress": pct})
    return render("causes.html", user=user, causes=causes)


@router.post("/causes/new")
async def cause_create(name: str = Form(...), db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    db.add(ContributionCause(name=name.strip()))
    await db.commit()
    return RedirectResponse(url="/causes", status_code=302)


# ── Contributions ──
@router.get("/contributions", response_class=HTMLResponse)
async def contribution_list(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    result = await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(desc(Contribution.date_paid)).limit(200))
    return render("contributions.html", user=user, contributions=result.scalars().all())


# ── Export ──
@router.get("/export/csv")
async def export_csv(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    import csv, io
    result = await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(Contribution.date_paid))
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Member #", "Member Name", "Cause", "Amount (KES)", "Date Paid", "Notes"])
    for c in result.scalars().all():
        w.writerow([c.member.member_number, c.member.name, c.cause.name, float(c.amount), c.date_paid.isoformat(), c.notes])
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=kh07_contributions.csv"})


@router.get("/export/excel")
async def export_excel(request: Request, db: AsyncSession = Depends(get_db), user: str = Depends(require_auth)):
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    def sh(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill, cell.border = h_font, h_fill, thin
    
    def aw(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 50)
    
    # Sheet 1
    ws = wb.active; ws.title = "Contributions"
    sh(ws, ["Member #", "Member Name", "Cause", "Amount (KES)", "Date Paid", "Notes"])
    for i, c in enumerate((await db.execute(select(Contribution).options(selectinload(Contribution.member), selectinload(Contribution.cause)).order_by(Contribution.date_paid))).scalars().all(), 2):
        ws.cell(row=i, column=1, value=c.member.member_number).border = thin
        ws.cell(row=i, column=2, value=c.member.name).border = thin
        ws.cell(row=i, column=3, value=c.cause.name).border = thin
        amt = ws.cell(row=i, column=4, value=float(c.amount)); amt.number_format = '#,##0'; amt.border = thin
        ws.cell(row=i, column=5, value=c.date_paid.isoformat()).border = thin
        ws.cell(row=i, column=6, value=c.notes).border = thin
    aw(ws)
    
    # Sheet 2
    ws2 = wb.create_sheet("Member Summary")
    sh(ws2, ["#", "Name", "Phone", "Total (KES)", "Contributions", "Active"])
    for i, m in enumerate((await db.execute(select(Member).order_by(Member.member_number))).scalars().all(), 2):
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.member_id == m.id))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.id)).where(Contribution.member_id == m.id))).scalar() or 0
        ws2.cell(row=i, column=1, value=m.member_number).border = thin
        ws2.cell(row=i, column=2, value=m.name).border = thin
        ws2.cell(row=i, column=3, value=m.phone_number).border = thin
        a = ws2.cell(row=i, column=4, value=total); a.number_format = '#,##0'; a.border = thin
        ws2.cell(row=i, column=5, value=count).border = thin
        ws2.cell(row=i, column=6, value="Yes" if m.is_active else "No").border = thin
    aw(ws2)
    
    # Sheet 3
    ws3 = wb.create_sheet("Cause Summary")
    sh(ws3, ["Cause", "Total (KES)", "Contributors", "Target (KES)", "Progress (%)"])
    for i, (cid, cname, tgt) in enumerate((await db.execute(select(ContributionCause.id, ContributionCause.name, ContributionCause.target_amount))), 2):
        total = float((await db.execute(select(func.coalesce(func.sum(Contribution.amount), 0)).where(Contribution.cause_id == cid))).scalar() or 0)
        count = (await db.execute(select(func.count(Contribution.member_id.distinct())).where(Contribution.cause_id == cid))).scalar() or 0
        ws3.cell(row=i, column=1, value=cname).border = thin
        a = ws3.cell(row=i, column=2, value=total); a.number_format = '#,##0'; a.border = thin
        ws3.cell(row=i, column=3, value=count).border = thin
        t = ws3.cell(row=i, column=4, value=float(tgt) if tgt else 0); t.number_format = '#,##0'; t.border = thin
        ws3.cell(row=i, column=5, value=round(float(total) / float(tgt) * 100, 1) if tgt and tgt > 0 else 0).border = thin
    aw(ws3)
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=kh07_contributions.xlsx"})
